import datetime
from tkinter import filedialog, messagebox

from openpyxl import load_workbook

import products
import state

HEADER_ALIASES = {
    "nome": {"nome", "denumire", "name", "product", "produs"},
    "quantita": {"quantita", "cantitate", "quantity", "qty", "stoc"},
    "prezzo": {"prezzo", "pret", "preț", "price"},
    "barcode": {"barcode", "cod", "codice", "ean"},
}


def _resolve_headers(header_row):
    mapping = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases:
                mapping[field] = idx
                break
    return mapping


def importa_excel():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")],
        title="Alege fișierul Excel",
    )
    if not file_path:
        return

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        messagebox.showerror("Eroare", f"Nu pot deschide fișierul:\n{e}")
        return

    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        messagebox.showwarning("Gol", "Fișierul e gol.")
        return

    cols = _resolve_headers(header)
    if not {"nome", "quantita", "prezzo"}.issubset(cols):
        # Fallback to positional: nome, quantita, prezzo, barcode
        cols = {"nome": 0, "quantita": 1, "prezzo": 2, "barcode": 3}
        # The first row was probably data, not a header — rewind by chaining it back.
        rows = _prepend(header, rows)

    now = str(datetime.datetime.now())
    added = 0
    updated = 0
    skipped = 0

    for raw in rows:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        try:
            nome = _get(raw, cols, "nome")
            quantita = int(float(_get(raw, cols, "quantita")))
            prezzo = float(_get(raw, cols, "prezzo"))
            barcode = _get(raw, cols, "barcode")
            barcode = str(barcode).strip() if barcode not in (None, "") else None
        except (TypeError, ValueError):
            skipped += 1
            continue

        if not nome or quantita is None:
            skipped += 1
            continue
        nome = str(nome).strip()

        existing = None
        if barcode:
            state.cursor.execute(
                "SELECT id, quantita FROM prodotti WHERE barcode = ?", (barcode,)
            )
            existing = state.cursor.fetchone()

        if existing:
            idp, stock = existing
            state.cursor.execute(
                "UPDATE prodotti SET quantita = quantita + ?, nome = ?, prezzo = ? WHERE id = ?",
                (quantita, nome, prezzo, idp),
            )
            state.cursor.execute(
                "INSERT INTO movimenti (id_prodotto, nome, tipo, quantita, data) VALUES (?, ?, ?, ?, ?)",
                (idp, nome, "încarcare", quantita, now),
            )
            updated += 1
        else:
            state.cursor.execute(
                "INSERT INTO prodotti (nome, quantita, prezzo, barcode) VALUES (?, ?, ?, ?)",
                (nome, quantita, prezzo, barcode),
            )
            idp = state.cursor.lastrowid
            state.cursor.execute(
                "INSERT INTO movimenti (id_prodotto, nome, tipo, quantita, data) VALUES (?, ?, ?, ?, ?)",
                (idp, nome, "încarcare", quantita, now),
            )
            added += 1

    state.conn.commit()
    products.aggiorna_tabella()

    messagebox.showinfo(
        "Import terminat",
        f"Adăugate: {added}\nActualizate: {updated}\nIgnorate: {skipped}",
    )


def _get(row, cols, field):
    idx = cols.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _prepend(first, rest):
    yield first
    for r in rest:
        yield r
